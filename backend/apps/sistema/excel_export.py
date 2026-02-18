"""
Utilidades para exportación de datos a Excel
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime


class ExcelExporter:
    """Helper para exportar datos a Excel con formato"""
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        
        # Estilos predefinidos
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.header_font = Font(bold=True, color='FFFFFF', size=12)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def exportar_ventas(self, ventas):
        """
        Exporta un listado de ventas a Excel
        
        Args:
            ventas: QuerySet de ventas
        
        Returns:
            HttpResponse con el archivo Excel
        """
        self.ws.title = 'Ventas'
        
        # Headers
        headers = [
            'Número', 'Fecha', 'Cliente', 'Total', 'Descuento',
            'Método de Pago', 'Estado', 'Usuario', 'Observaciones'
        ]
        self.ws.append(headers)
        
        # Aplicar formato a headers
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Datos
        for venta in ventas:
            self.ws.append([
                venta.numero,
                venta.fecha.strftime('%d/%m/%Y %H:%M'),
                venta.cliente.nombre_completo,
                float(venta.total),
                float(venta.descuento_monto),
                venta.metodo_pago,
                venta.estado,
                venta.usuario.username,
                (venta.motivo_anulacion or '') if venta.estado == 'ANULADA' else ''
            ])
        
        # Ajustar anchos de columna
        self._auto_adjust_columns()
        
        # Retornar como respuesta HTTP
        return self._generate_response('ventas')
    
    def exportar_productos(self, variantes):
        """
        Exporta un listado de productos (variantes) a Excel
        
        Args:
            variantes: QuerySet de variantes de producto
        
        Returns:
            HttpResponse con el archivo Excel
        """
        self.ws.title = 'Productos'
        
        # Headers
        headers = [
            'SKU', 'Código de Barras', 'Producto', 'Variante',
            'Costo', 'Precio Contado', 'Precio Tarjeta', 'Stock Total', 'Estado'
        ]
        self.ws.append(headers)
        
        # Formato headers
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Datos
        for variante in variantes:
            self.ws.append([
                variante.sku,
                variante.codigo_barras or '',
                variante.producto.nombre,
                variante.nombre_completo,
                float(variante.costo),
                float(variante.precio_mostrador),
                float(variante.precio_tarjeta),
                variante.stock_total,
                variante.estado
            ])
        
        self._auto_adjust_columns()
        return self._generate_response('productos')
    
    def exportar_clientes(self, clientes):
        """
        Exporta un listado de clientes a Excel
        
        Args:
            clientes: QuerySet de clientes
        
        Returns:
            HttpResponse con el archivo Excel
        """
        self.ws.title = 'Clientes'
        
        headers = [
            'DNI/CUIT', 'Nombre Completo', 'Email', 'Teléfono',
            'Dirección', 'Localidad', 'Provincia', 'Tipo'
        ]
        self.ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        for cliente in clientes:
            self.ws.append([
                cliente.dni_cuit,
                cliente.nombre_completo,
                cliente.email or '',
                cliente.telefono or '',
                cliente.direccion or '',
                cliente.localidad or '',
                cliente.provincia or '',
                cliente.tipo_cliente
            ])
        
        self._auto_adjust_columns()
        return self._generate_response('clientes')
    
    def exportar_inventario(self, stocks):
        """
        Exporta inventario (stocks por depósito) a Excel
        
        Args:
            stocks: QuerySet de Stock
        
        Returns:
            HttpResponse con el archivo Excel
        """
        self.ws.title = 'Inventario'
        
        headers = [
            'Producto', 'SKU', 'Depósito', 'Cantidad',
            'Stock Mínimo', 'Stock Máximo', 'Alerta'
        ]
        self.ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        for stock in stocks:
            alerta = 'CRÍTICO' if stock.esta_critico else ('BAJO' if stock.esta_bajo else 'OK')
            self.ws.append([
                stock.variante.nombre_completo,
                stock.variante.sku,
                stock.deposito.nombre,
                stock.cantidad,
                stock.stock_minimo,
                stock.stock_maximo,
                alerta
            ])
        
        self._auto_adjust_columns()
        return self._generate_response('inventario')
    
    def exportar_compras(self, compras):
        """Exporta un listado de compras a Excel"""
        self.ws.title = 'Compras'
        
        headers = [
            'Número', 'Fecha', 'Proveedor', 'Total',
            'Estado', 'Usuario', 'Observaciones'
        ]
        self.ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        for compra in compras:
            self.ws.append([
                compra.numero,
                compra.fecha.strftime('%d/%m/%Y %H:%M'),
                compra.proveedor.nombre,
                float(compra.total),
                compra.estado,
                compra.usuario.username,
                compra.observaciones or ''
            ])
        
        self._auto_adjust_columns()
        return self._generate_response('compras')

    def exportar_reporte_productos_mas_vendidos(self, filas):
        """
        Exporta reporte de productos más vendidos.
        filas: lista de dicts con keys: variante__sku, variante__producto_base__nombre o nombre, cantidad_vendida, total_facturado
        """
        self.ws.title = 'Productos más vendidos'
        headers = ['SKU', 'Producto', 'Cantidad vendida', 'Total facturado']
        self.ws.append(headers)
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        for row in filas:
            nombre = row.get('variante__producto_base__nombre') or row.get('nombre') or ''
            sku = row.get('variante__sku') or row.get('sku') or ''
            self.ws.append([
                sku,
                nombre,
                int(row.get('cantidad_vendida', 0)),
                float(row.get('total_facturado', 0)),
            ])
        self._auto_adjust_columns()
        return self._generate_response('productos_mas_vendidos')

    def _auto_adjust_columns(self):
        """Ajusta automáticamente el ancho de las columnas"""
        for column_cells in self.ws.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            self.ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)
    
    def _generate_response(self, filename_base):
        """
        Genera un HttpResponse con el archivo Excel
        
        Args:
            filename_base: Nombre base del archivo
        
        Returns:
            HttpResponse
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'{filename_base}_{timestamp}.xlsx'
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        self.wb.save(response)
        return response

"""
API pública de la tienda web (avila-web).
Endpoints sin autenticación para catálogo y pedidos.
Endpoints con auth para administrador.
"""
import logging
from decimal import Decimal
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.db import models as db_models
from django.core.paginator import Paginator
from django.conf import settings
from django.core.mail import send_mail
from django.template.loader import render_to_string

from apps.productos.models import VarianteProducto, Marca, Categoria, ProductoBase
from apps.inventario.models import Deposito, Stock
from apps.inventario.services import InventarioService
from apps.inventario.models import MovimientoStock
from apps.ventas.models import Venta, DetalleVenta
from apps.usuarios.models import Usuario
from .models import PuntoRetiro, ClienteWeb, PedidoWeb, ModeloMoto
from .mercadopago import crear_preferencia_para_venta
from .auth import (
    hash_password, verify_password,
    generate_token, get_cliente_from_request,
    verify_google_token,
)

logger = logging.getLogger('apps')

_METODO_LABELS = {
    'MERCADOPAGO':   'Mercado Pago',
    'TRANSFERENCIA': 'Transferencia bancaria',
    'EFECTIVO':      'Efectivo',
    'TARJETA':       'Tarjeta',
}
_ESTADO_LABELS = {
    'PENDIENTE_PAGO':  'Pendiente de pago',
    'PAGO_CONFIRMADO': 'Pago confirmado',
    'EN_PREPARACION':  'En preparación',
    'ENVIADO':         'Enviado',
    'ENTREGADO':       'Entregado',
    'COMPLETADA':      'Completada',
    'ANULADA':         'Anulada',
}


def _enviar_emails_pedido(venta, datos_cliente, items_para_venta, tipo_entrega, direccion_envio):
    """Envía email de confirmación al cliente y aviso al local."""
    from urllib.parse import quote

    items_ctx = []
    for it in items_para_venta:
        sub = (it['precio_unitario'] - it['descuento_unitario']) * it['cantidad']
        items_ctx.append({
            'nombre':   it['variante'].nombre_completo,
            'codigo':   it['variante'].codigo,
            'cantidad': it['cantidad'],
            'precio':   f'{it["precio_unitario"]:,.2f}',
            'subtotal': f'{sub:,.2f}',
        })

    total_fmt = f'{venta.total:,.2f}'
    nombre_cliente = datos_cliente.get('nombre', '')
    email_cliente  = datos_cliente.get('email', '')
    telefono       = datos_cliente.get('telefono', '')
    metodo         = venta.metodo_pago
    wa_text        = quote(f'Hola! Quiero consultar mi pedido #{venta.numero}')

    tipo_label = 'Retiro en local' if tipo_entrega == 'retiro' else 'Envío a domicilio'

    ctx_cliente = {
        'venta_numero':    venta.numero,
        'nombre_cliente':  nombre_cliente,
        'items':           items_ctx,
        'total':           total_fmt,
        'metodo_pago':     metodo,
        'tipo_entrega':    tipo_entrega,
        'direccion_envio': direccion_envio,
        'bank_titular':    'Avila Marcelo Bernabe',
        'bank_alias':      'avilaxxx',
        'bank_banco':      'Mercado Pago',
        'bank_cbu':        '',
        'wa_text':         wa_text,
    }
    ctx_local = {
        'venta_numero':      venta.numero,
        'nombre_cliente':    nombre_cliente,
        'email_cliente':     email_cliente,
        'telefono_cliente':  telefono,
        'items':             items_ctx,
        'total':             total_fmt,
        'metodo_pago_label': _METODO_LABELS.get(metodo, metodo),
        'tipo_entrega_label': tipo_label,
        'direccion_envio':   direccion_envio,
    }

    email_local = getattr(settings, 'EMAIL_LOCAL', '') or getattr(settings, 'EMAIL_HOST_USER', '')

    try:
        if email_cliente:
            html_cliente = render_to_string('emails/confirmacion_pedido.html', ctx_cliente)
            send_mail(
                subject=f'Pedido #{venta.numero} recibido — Avila Moto Repuestos',
                message=f'Tu pedido #{venta.numero} fue recibido. Total: ${total_fmt}',
                from_email=settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else email_local,
                recipient_list=[email_cliente],
                html_message=html_cliente,
                fail_silently=True,
            )
            logger.info(f'[Email] Confirmación enviada a {email_cliente} para pedido #{venta.numero}')
    except Exception as e:
        logger.warning(f'[Email] Error enviando confirmación al cliente: {e}')

    try:
        if email_local:
            html_local = render_to_string('emails/aviso_pedido_local.html', ctx_local)
            send_mail(
                subject=f'🛒 Nuevo pedido web #{venta.numero} — ${total_fmt}',
                message=f'Nuevo pedido web #{venta.numero} de {nombre_cliente}. Total: ${total_fmt}',
                from_email=settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else email_local,
                recipient_list=[email_local],
                html_message=html_local,
                fail_silently=True,
            )
            logger.info(f'[Email] Aviso de nuevo pedido enviado a {email_local}')
    except Exception as e:
        logger.warning(f'[Email] Error enviando aviso al local: {e}')


def _enviar_email_estado(venta, email_cliente, nombre_cliente):
    """Notifica al cliente cuando cambia el estado de su pedido."""
    if not email_cliente:
        return
    estado_label = _ESTADO_LABELS.get(venta.estado, venta.estado)
    try:
        send_mail(
            subject=f'Tu pedido #{venta.numero} — {estado_label}',
            message=(
                f'Hola {nombre_cliente},\n\n'
                f'El estado de tu pedido #{venta.numero} cambió a: {estado_label}.\n\n'
                f'Para consultas: https://wa.me/5493834625390\n\n'
                f'Avila Moto Repuestos\nAv. Pte. Castillo 1165, Catamarca'
            ),
            from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', getattr(settings, 'EMAIL_HOST_USER', '')),
            recipient_list=[email_cliente],
            fail_silently=True,
        )
        logger.info(f'[Email] Estado {venta.estado} notificado a {email_cliente} para pedido #{venta.numero}')
    except Exception as e:
        logger.warning(f'[Email] Error notificando estado: {e}')


def _get_deposito_principal():
    """Depósito principal para stock web."""
    return Deposito.objects.filter(es_principal=True, activo=True).first()


def _get_usuario_venta_web():
    """Usuario para registrar ventas web (primer admin)."""
    return Usuario.objects.filter(is_active=True).filter(
        db_models.Q(is_superuser=True) | db_models.Q(rol=Usuario.Rol.ADMINISTRADOR)
    ).order_by('id').first()


def _get_imagen_url(producto_base, request=None):
    """Devuelve la URL de la imagen principal del producto (galería o campo legacy)."""
    if not producto_base:
        return None
    img = producto_base.imagenes.filter(es_principal=True).first() or producto_base.imagenes.first()
    if img and img.imagen:
        return request.build_absolute_uri(img.imagen.url) if request else img.imagen.url
    if producto_base.imagen:
        return request.build_absolute_uri(producto_base.imagen.url) if request else producto_base.imagen.url
    return None


def _variante_a_dict(v, stock_cantidad=0, incluir_marca_cat=True, request=None):
    """Convierte variante a dict para API tienda."""
    data = {
        'id': v.id,
        'codigo': v.codigo,
        'nombre_completo': v.nombre_completo,
        'precio_web': str(v.precio_web),
        'stock': stock_cantidad,
    }
    if incluir_marca_cat:
        pb = v.producto_base
        data['marca'] = pb.marca.nombre if pb else ''
        data['marca_id'] = pb.marca_id if pb else None
        data['categoria'] = pb.categoria.nombre if pb else ''
        data['categoria_id'] = pb.categoria_id if pb else None
        data['descripcion'] = (pb.descripcion or '')[:300] if pb else ''
        data['imagen_url'] = _get_imagen_url(pb, request)
        # Motos compatibles (ids para filtros frontend)
        if pb:
            data['motos_compatibles'] = list(
                pb.modelos_compatibles.filter(activo=True).values('id', 'marca', 'modelo', 'anio')
            )
        else:
            data['motos_compatibles'] = []
    return data


@api_view(['GET'])
@permission_classes([AllowAny])
def productos_list(request):
    """
    Lista variantes con stock y precio_web (público).
    Query: page, page_size, categoria, marca, search
    """
    deposito = _get_deposito_principal()
    if not deposito:
        return Response(
            {'error': 'No hay depósito principal configurado'},
            status=status.HTTP_503_SERVICE_UNAVAILABLE
        )

    from apps.productos.models import ProductoImagen
    qs = VarianteProducto.objects.filter(
        activo=True,
        producto_base__activo=True,
        precio_web__gt=0,
    ).select_related('producto_base', 'producto_base__marca', 'producto_base__categoria').prefetch_related(
        db_models.Prefetch('stocks', queryset=Stock.objects.filter(deposito=deposito)),
        db_models.Prefetch('producto_base__imagenes', queryset=ProductoImagen.objects.order_by('orden', 'id')),
        'producto_base__modelos_compatibles',
    )

    categoria = request.query_params.get('categoria')
    if categoria:
        qs = qs.filter(producto_base__categoria_id=categoria)

    marca = request.query_params.get('marca')
    if marca:
        qs = qs.filter(producto_base__marca_id=marca)

    search = request.query_params.get('search', '').strip()
    if search:
        from django.db.models import Q
        qs = qs.filter(
            Q(codigo__icontains=search) |
            Q(nombre_variante__icontains=search) |
            Q(producto_base__nombre__icontains=search) |
            Q(producto_base__marca__nombre__icontains=search)
        )

    modelo_id = request.query_params.get('modelo')
    if modelo_id:
        qs = qs.filter(producto_base__modelos_compatibles__id=modelo_id)

    _ORDERING = {
        'nombre': ('producto_base__nombre', 'nombre_variante'),
        '-nombre': ('-producto_base__nombre', '-nombre_variante'),
        'precio_web': ('precio_web',),
        '-precio_web': ('-precio_web',),
    }
    ordering_param = request.query_params.get('ordering', '').strip()
    qs = qs.order_by(*_ORDERING.get(ordering_param, ('producto_base__nombre', 'nombre_variante')))

    page_size = min(int(request.query_params.get('page_size', 24)), 100)
    paginator = Paginator(qs, page_size)
    page_num = request.query_params.get('page', 1)
    page = paginator.get_page(page_num)

    items = []
    for v in page.object_list:
        stock_qs = v.stocks.filter(deposito=deposito)
        cantidad = stock_qs.first().cantidad if stock_qs.exists() else 0
        items.append(_variante_a_dict(v, cantidad, request=request))

    return Response({
        'count': paginator.count,
        'total_pages': paginator.num_pages,
        'current_page': page.number,
        'results': items,
    })


@api_view(['GET'])
@permission_classes([AllowAny])
def producto_detail(request, pk):
    """Detalle de una variante (público)."""
    deposito = _get_deposito_principal()
    from apps.productos.models import ProductoImagen
    try:
        v = VarianteProducto.objects.select_related(
            'producto_base', 'producto_base__marca', 'producto_base__categoria'
        ).prefetch_related(
            db_models.Prefetch(
                'stocks',
                queryset=Stock.objects.filter(deposito=deposito) if deposito else Stock.objects.none()
            ),
            db_models.Prefetch('producto_base__imagenes', queryset=ProductoImagen.objects.order_by('orden', 'id')),
            'producto_base__modelos_compatibles',
        ).get(pk=pk, activo=True, producto_base__activo=True)
    except VarianteProducto.DoesNotExist:
        return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    stock_qs = v.stocks.filter(deposito=deposito)
    cantidad = stock_qs.first().cantidad if stock_qs.exists() else 0
    data = _variante_a_dict(v, cantidad, request=request)
    data['descripcion'] = v.producto_base.descripcion or ''
    # Todas las imágenes de la galería
    data['imagenes'] = [
        {'id': img.id, 'url': request.build_absolute_uri(img.imagen.url) if img.imagen else None, 'es_principal': img.es_principal}
        for img in v.producto_base.imagenes.all()
    ]
    return Response(data)


@api_view(['GET'])
@permission_classes([AllowAny])
def categorias_list(request):
    """Lista categorías activas (público)."""
    cats = Categoria.objects.filter(activo=True).order_by('nombre').values('id', 'nombre')
    return Response(list(cats))


@api_view(['GET'])
@permission_classes([AllowAny])
def marcas_list(request):
    """Lista marcas activas (público)."""
    marcas = Marca.objects.filter(activo=True).order_by('nombre').values('id', 'nombre')
    return Response(list(marcas))


@api_view(['GET'])
@permission_classes([AllowAny])
def puntos_retiro_list(request):
    """Lista puntos de retiro activos para la tienda web."""
    puntos = PuntoRetiro.objects.filter(activo=True).order_by('nombre')
    data = [
        {
            'id': p.id,
            'nombre': p.nombre,
            'direccion_texto': p.direccion_texto or '',
            'lat': p.lat,
            'lng': p.lng,
            'telefono': p.telefono or '',
            'horarios': p.horarios or '',
        }
        for p in puntos
    ]
    return Response(data)


@api_view(['GET'])
@permission_classes([AllowAny])
def modelos_moto_list(request):
    """Lista de modelos de moto activos para el selector de compatibilidad."""
    modelos = ModeloMoto.objects.filter(activo=True).values('id', 'marca', 'modelo', 'anio')
    return Response(list(modelos))


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def productos_por_moto(request, moto_id):
    """
    Devuelve los productos compatibles con una moto específica (para el POS).
    Incluye variantes con stock_actual.
    """
    try:
        moto = ModeloMoto.objects.get(id=moto_id, activo=True)
    except ModeloMoto.DoesNotExist:
        return Response({'error': 'Moto no encontrada'}, status=status.HTTP_404_NOT_FOUND)

    from apps.productos.models import VarianteProducto
    from apps.inventario.models import Stock
    from django.db.models import Sum

    productos_base = moto.productos.filter(activo=True).prefetch_related('variantes', 'marca')
    data = []
    for pb in productos_base:
        for variante in pb.variantes.filter(activo=True):
            stock_actual = Stock.objects.filter(
                variante=variante, deposito__activo=True
            ).aggregate(total=Sum('cantidad'))['total'] or 0
            data.append({
                'id': variante.id,
                'codigo': variante.codigo,
                'nombre_completo': variante.nombre_completo,
                'nombre_variante': variante.nombre_variante,
                'marca': pb.marca.nombre,
                'precio_mostrador': float(variante.precio_mostrador),
                'precio_tarjeta': float(variante.precio_tarjeta),
                'stock_actual': stock_actual,
            })

    return Response({'moto': str(moto), 'count': len(data), 'results': data})


# ── Gestión de modelos de moto (admin POS) ────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def motos_crud(request):
    """GET: lista todos. POST: crea nuevo modelo de moto."""
    if request.method == 'GET':
        qs = ModeloMoto.objects.order_by('marca', 'modelo', 'anio')
        activo = request.query_params.get('activo')
        if activo is not None:
            qs = qs.filter(activo=activo.lower() in ('true', '1'))
        data = [{'id': m.id, 'marca': m.marca, 'modelo': m.modelo, 'anio': m.anio, 'activo': m.activo} for m in qs]
        return Response(data)
    # POST
    d = request.data
    marca = (d.get('marca') or '').strip()
    modelo = (d.get('modelo') or '').strip()
    try:
        anio = int(d.get('anio') or 0)
    except (ValueError, TypeError):
        return Response({'error': 'Año inválido'}, status=400)
    if not marca or not modelo or not anio:
        return Response({'error': 'marca, modelo y anio son requeridos'}, status=400)
    obj, created = ModeloMoto.objects.get_or_create(
        marca=marca, modelo=modelo, anio=anio,
        defaults={'activo': True}
    )
    return Response({'id': obj.id, 'marca': obj.marca, 'modelo': obj.modelo, 'anio': obj.anio, 'activo': obj.activo, 'created': created},
                    status=201 if created else 200)


@api_view(['PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def moto_detail(request, moto_id):
    """PATCH: actualiza. DELETE: elimina."""
    try:
        moto = ModeloMoto.objects.get(id=moto_id)
    except ModeloMoto.DoesNotExist:
        return Response({'error': 'No encontrado'}, status=404)
    if request.method == 'DELETE':
        moto.delete()
        return Response(status=204)
    # PATCH
    d = request.data
    if 'marca' in d:
        moto.marca = d['marca'].strip()
    if 'modelo' in d:
        moto.modelo = d['modelo'].strip()
    if 'anio' in d:
        moto.anio = int(d['anio'])
    if 'activo' in d:
        moto.activo = bool(d['activo'])
    moto.save()
    return Response({'id': moto.id, 'marca': moto.marca, 'modelo': moto.modelo, 'anio': moto.anio, 'activo': moto.activo})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def motos_importar_excel(request):
    """
    Importa modelos de moto desde un Excel.
    Detecta columnas automáticamente — acepta variantes de nombres:
      marca  → marca, brand
      modelo → modelo, model, nombre
      año    → año, anio, year, anio_modelo, año_modelo, anio_moto
    """
    import openpyxl
    import unicodedata

    archivo = request.FILES.get('file')
    if not archivo:
        return Response({'error': 'Enviá un archivo Excel en el campo "file"'}, status=400)

    def normalizar_header(v):
        """Minúsculas, sin tildes, sin espacios extras."""
        s = str(v or '').strip().lower()
        s = unicodedata.normalize('NFD', s)
        s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        return s

    ALIAS_MARCA  = {'marca', 'brand', 'fabricante'}
    ALIAS_MODELO = {'modelo', 'model', 'nombre_modelo', 'nombre'}
    ALIAS_ANIO   = {'anio', 'ano', 'year', 'anio_modelo', 'ano_modelo',
                    'anio_moto', 'ano_moto', 'año', 'año_modelo', 'año_moto'}

    def find_col(headers_norm, aliases):
        """Devuelve el índice de la primera columna cuyo header está en aliases."""
        for idx, h in enumerate(headers_norm):
            if h in aliases:
                return idx
        # Segunda pasada: buscar si algún alias es substring del header
        for idx, h in enumerate(headers_norm):
            for alias in aliases:
                if alias in h and len(alias) >= 3:
                    return idx
        return None

    def cell_val(row, idx):
        if idx is None or idx >= len(row):
            return ''
        v = row[idx].value
        return str(v).strip() if v is not None else ''

    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
        ws = wb.active

        raw_headers = [c.value for c in ws[1]]
        headers_norm = [normalizar_header(h) for h in raw_headers]

        idx_marca  = find_col(headers_norm, ALIAS_MARCA)
        idx_modelo = find_col(headers_norm, ALIAS_MODELO)
        idx_anio   = find_col(headers_norm, ALIAS_ANIO)

        if idx_marca is None or idx_modelo is None or idx_anio is None:
            faltantes = []
            if idx_marca is None:  faltantes.append('marca')
            if idx_modelo is None: faltantes.append('modelo')
            if idx_anio is None:   faltantes.append('año/anio')
            return Response({
                'error': f'No se encontraron las columnas: {", ".join(faltantes)}. '
                         f'Encabezados detectados: {[h for h in raw_headers if h]}',
            }, status=400)

        creados, existentes, errores = 0, 0, []
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            marca  = cell_val(row, idx_marca)
            modelo = cell_val(row, idx_modelo)
            anio_str = cell_val(row, idx_anio)

            if not marca or not modelo:
                continue
            try:
                anio = int(float(anio_str)) if anio_str else 0
                if not anio:
                    raise ValueError('Año vacío o no numérico')
                obj, created = ModeloMoto.objects.get_or_create(
                    marca=marca, modelo=modelo, anio=anio,
                    defaults={'activo': True}
                )
                creados    += int(created)
                existentes += int(not created)
            except Exception as e:
                errores.append({'fila': i, 'marca': marca, 'modelo': modelo, 'anio': anio_str, 'error': str(e)})

        return Response({
            'creados': creados,
            'existentes': existentes,
            'errores': errores[:20],  # máximo 20 errores en la respuesta
            'total_errores': len(errores),
            'columnas_detectadas': {
                'marca':  raw_headers[idx_marca],
                'modelo': raw_headers[idx_modelo],
                'anio':   raw_headers[idx_anio],
            },
        })
    except Exception as e:
        return Response({'error': str(e)}, status=400)


# ── Asignación de compatibilidad (admin POS) ──────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def producto_motos_compat(request, producto_base_id):
    """Lista las motos compatibles de un ProductoBase."""
    from apps.productos.models import ProductoBase
    try:
        pb = ProductoBase.objects.get(id=producto_base_id)
    except ProductoBase.DoesNotExist:
        return Response({'error': 'Producto no encontrado'}, status=404)
    motos = pb.modelos_compatibles.filter(activo=True).order_by('marca', 'modelo', 'anio')
    return Response([{'id': m.id, 'marca': m.marca, 'modelo': m.modelo, 'anio': m.anio} for m in motos])


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def asignar_moto_producto(request, producto_base_id):
    """Agrega una moto compatible a un ProductoBase. Body: {moto_id}"""
    from apps.productos.models import ProductoBase
    try:
        pb = ProductoBase.objects.get(id=producto_base_id)
    except ProductoBase.DoesNotExist:
        return Response({'error': 'Producto no encontrado'}, status=404)
    moto_id = request.data.get('moto_id')
    if not moto_id:
        return Response({'error': 'moto_id es requerido'}, status=400)
    try:
        moto = ModeloMoto.objects.get(id=moto_id)
    except ModeloMoto.DoesNotExist:
        return Response({'error': 'Moto no encontrada'}, status=404)
    pb.modelos_compatibles.add(moto)
    return Response({'ok': True, 'moto': {'id': moto.id, 'marca': moto.marca, 'modelo': moto.modelo, 'anio': moto.anio}})


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def quitar_moto_producto(request, producto_base_id, moto_id):
    """Quita una moto compatible de un ProductoBase."""
    from apps.productos.models import ProductoBase
    try:
        pb = ProductoBase.objects.get(id=producto_base_id)
        moto = ModeloMoto.objects.get(id=moto_id)
    except (ProductoBase.DoesNotExist, ModeloMoto.DoesNotExist):
        return Response({'error': 'No encontrado'}, status=404)
    pb.modelos_compatibles.remove(moto)
    return Response(status=204)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def asignar_moto_masivo(request):
    """
    Asigna una moto a múltiples productos base de una vez.
    Body: {moto_id, producto_base_ids: [1,2,3,...]}
    """
    from apps.productos.models import ProductoBase
    moto_id = request.data.get('moto_id')
    ids = request.data.get('producto_base_ids', [])
    if not moto_id or not ids:
        return Response({'error': 'moto_id y producto_base_ids son requeridos'}, status=400)
    try:
        moto = ModeloMoto.objects.get(id=moto_id)
    except ModeloMoto.DoesNotExist:
        return Response({'error': 'Moto no encontrada'}, status=404)
    productos = ProductoBase.objects.filter(id__in=ids)
    for pb in productos:
        pb.modelos_compatibles.add(moto)
    return Response({'ok': True, 'asignados': productos.count(), 'moto': str(moto)})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def moto_toggle_activo(request, moto_id):
    """Activa/desactiva un modelo de moto."""
    try:
        moto = ModeloMoto.objects.get(id=moto_id)
    except ModeloMoto.DoesNotExist:
        return Response({'error': 'No encontrado'}, status=404)
    moto.activo = not moto.activo
    moto.save()
    return Response({'id': moto.id, 'activo': moto.activo})


@api_view(['POST'])
@permission_classes([AllowAny])
def mercadopago_crear_preferencia(request):
    """
    Crea una preferencia de pago de Mercado Pago para una venta web existente.
    Body: { venta_id: number, back_urls?: {success, pending, failure}, auto_return?: 'approved' }
    """
    data = request.data or {}
    venta_id = data.get('venta_id')
    if not venta_id:
        return Response({'error': 'venta_id es obligatorio'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        venta = Venta.objects.get(pk=venta_id)
    except Venta.DoesNotExist:
        return Response({'error': 'Venta no encontrada'}, status=status.HTTP_404_NOT_FOUND)

    # Verificar que la venta tenga movimientos de tipo VENTA_WEB (pedido web)
    es_web = MovimientoStock.objects.filter(
        referencia_tipo='venta',
        referencia_id=venta.id,
        tipo=MovimientoStock.TipoMovimiento.VENTA_WEB,
    ).exists()
    if not es_web:
        return Response({'error': 'La venta no corresponde a un pedido web'}, status=status.HTTP_400_BAD_REQUEST)

    back_urls = data.get('back_urls') or {}
    auto_return = data.get('auto_return') or None

    pref = crear_preferencia_para_venta(venta, back_urls=back_urls, auto_return=auto_return)
    return Response(pref, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([AllowAny])
def mercadopago_webhook(request):
    """
    Webhook de Mercado Pago.
    MP envía notificaciones cuando cambia el estado de un pago.
    Docs: https://www.mercadopago.com.ar/developers/es/docs/your-integrations/notifications/webhooks
    """
    import logging
    logger = logging.getLogger('apps')

    data = request.data or {}
    topic = data.get('topic') or request.query_params.get('topic', '')
    payment_id = data.get('data', {}).get('id') or request.query_params.get('id', '')

    logger.info(f'[MP Webhook] topic={topic} payment_id={payment_id} data={data}')

    if topic not in ('payment', 'merchant_order', ''):
        return Response({'ok': True}, status=status.HTTP_200_OK)

    if not payment_id:
        return Response({'ok': True}, status=status.HTTP_200_OK)

    token = getattr(__import__('django.conf', fromlist=['settings']).settings, 'MERCADOPAGO_ACCESS_TOKEN', '')
    if not token:
        logger.warning('[MP Webhook] Sin ACCESS_TOKEN configurado, ignorando notificación')
        return Response({'ok': True}, status=status.HTTP_200_OK)

    try:
        import requests as req
        resp = req.get(
            f'https://api.mercadopago.com/v1/payments/{payment_id}',
            headers={'Authorization': f'Bearer {token}'},
            timeout=10,
        )
        if resp.status_code != 200:
            logger.warning(f'[MP Webhook] No se pudo consultar payment {payment_id}: {resp.status_code}')
            return Response({'ok': True}, status=status.HTTP_200_OK)

        pago = resp.json()
        mp_status = pago.get('status', '')
        external_ref = pago.get('external_reference', '')

        logger.info(f'[MP Webhook] payment={payment_id} status={mp_status} ref={external_ref}')

        if not external_ref:
            return Response({'ok': True}, status=status.HTTP_200_OK)

        try:
            venta = Venta.objects.get(pk=int(external_ref))
        except (Venta.DoesNotExist, ValueError):
            return Response({'ok': True}, status=status.HTTP_200_OK)

        if mp_status == 'approved' and venta.estado == Venta.EstadoVenta.PENDIENTE_PAGO:
            venta.estado = Venta.EstadoVenta.PAGO_CONFIRMADO
            venta.save(update_fields=['estado'])
            logger.info(f'[MP Webhook] Venta #{venta.numero} → PAGO_CONFIRMADO')
        elif mp_status in ('rejected', 'cancelled') and venta.estado == Venta.EstadoVenta.PENDIENTE_PAGO:
            venta.estado = Venta.EstadoVenta.ANULADA
            venta.save(update_fields=['estado'])
            logger.info(f'[MP Webhook] Venta #{venta.numero} → ANULADA (MP rechazó)')

    except Exception as e:
        logger.error(f'[MP Webhook] Error procesando payment {payment_id}: {e}')

    return Response({'ok': True}, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([AllowAny])
def pedido_create(request):
    """
    Crea un pedido (venta web).
    Body: { line_items: [{ variante_id, cantidad }], datos_cliente?: {}, observaciones?: "" }
    """
    from django.db import transaction

    data = request.data or {}
    line_items = data.get('line_items') or []
    datos_cliente = data.get('datos_cliente') or {}
    observaciones = data.get('observaciones') or 'Pedido tienda web'

    # Información de entrega
    tipo_entrega = (data.get('tipo_entrega') or '').strip().lower()  # 'retiro' | 'envio' | ''
    punto_retiro = (data.get('punto_retiro') or '').strip()
    punto_retiro_id = data.get('punto_retiro_id')
    # Datos de envío simples (fase 2)
    direccion_envio = (datos_cliente.get('direccion') or '').strip()
    localidad_envio = (datos_cliente.get('localidad') or '').strip()
    cp_envio = (datos_cliente.get('cp') or '').strip()
    provincia_envio = (datos_cliente.get('provincia') or '').strip()
    # Coordenadas opcionales de entrega (mapa)
    lat_entrega = data.get('lat_entrega')
    lng_entrega = data.get('lng_entrega')
    # Costo de envío base simple (se puede mejorar luego)
    costo_envio = Decimal('0')

    if datos_cliente:
        cliente_str = datos_cliente.get('nombre', '') or ''
        if datos_cliente.get('email'):
            cliente_str += f" - {datos_cliente['email']}"
        if datos_cliente.get('telefono'):
            cliente_str += f" - Tel: {datos_cliente['telefono']}"
        if cliente_str.strip():
            observaciones = f"{observaciones} | Cliente: {cliente_str.strip()}"

    # Descripción de entrega en observaciones
    if tipo_entrega:
        if tipo_entrega not in ('retiro', 'envio'):
            return Response(
                {'error': "tipo_entrega inválido. Debe ser 'retiro' o 'envio'."},
                status=status.HTTP_400_BAD_REQUEST
            )
        if tipo_entrega == 'retiro':
            punto_retiro_obj = None
            if punto_retiro_id:
                try:
                    punto_retiro_obj = PuntoRetiro.objects.get(id=punto_retiro_id, activo=True)
                except PuntoRetiro.DoesNotExist:
                    return Response(
                        {'error': 'Punto de retiro inválido o inactivo.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            elif punto_retiro:
                # Compatibilidad hacia atrás: texto libre
                observaciones = f"{observaciones} | Entrega: RETIRO EN LOCAL - {punto_retiro}"
            else:
                return Response(
                    {'error': 'Para retiro en local debe especificar punto_retiro_id.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if punto_retiro_obj:
                desc = f"{punto_retiro_obj.nombre}"
                if punto_retiro_obj.direccion_texto:
                    desc += f" - {punto_retiro_obj.direccion_texto}"
                observaciones = f"{observaciones} | Entrega: RETIRO EN LOCAL - {desc}"
        elif tipo_entrega == 'envio':
            # Validaciones mínimas de dirección para envío
            if not direccion_envio or not localidad_envio:
                return Response(
                    {'error': 'Para envío a domicilio debe indicar al menos dirección y localidad.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            # Cálculo simple de costo de envío (ejemplo fijo)
            costo_envio = Decimal('0')
            if provincia_envio:
                # Ejemplo: si provincia distinta de vacío, costo base
                costo_envio = Decimal('1500')
            else:
                costo_envio = Decimal('1000')
            observaciones = (
                f"{observaciones} | Entrega: ENVÍO A DOMICILIO"
                f" - Dirección: {direccion_envio}, {localidad_envio}"
                f"{f' (CP {cp_envio})' if cp_envio else ''}"
                f"{f', {provincia_envio}' if provincia_envio else ''}"
                f" | Costo envío estimado: ${costo_envio}"
            )
            # Agregar coordenadas si se recibieron
            if lat_entrega is not None and lng_entrega is not None:
                try:
                    lat_f = float(lat_entrega)
                    lng_f = float(lng_entrega)
                    observaciones = f"{observaciones} | Ubicación mapa: lat={lat_f:.6f}, lng={lng_f:.6f}"
                except (TypeError, ValueError):
                    # Si vienen mal formateadas, simplemente las ignoramos
                    pass

    if not line_items:
        return Response(
            {'error': 'line_items es obligatorio y no puede estar vacío'},
            status=status.HTTP_400_BAD_REQUEST
        )

    deposito = _get_deposito_principal()
    if not deposito:
        return Response(
            {'error': 'No hay depósito principal configurado'},
            status=status.HTTP_503_SERVICE_UNAVAILABLE
        )

    usuario = _get_usuario_venta_web()
    if not usuario:
        return Response(
            {'error': 'No hay usuario para registrar ventas web'},
            status=status.HTTP_503_SERVICE_UNAVAILABLE
        )

    items_para_venta = []
    errores = []

    for idx, item in enumerate(line_items):
        cantidad = item.get('cantidad')
        if not cantidad or int(cantidad) < 1:
            errores.append(f'Item {idx + 1}: cantidad inválida')
            continue

        variante_id = item.get('variante_id')
        if not variante_id:
            errores.append(f'Item {idx + 1}: variante_id requerido')
            continue

        try:
            variante = VarianteProducto.objects.get(pk=variante_id, activo=True)
        except (VarianteProducto.DoesNotExist, ValueError):
            errores.append(f'Item {idx + 1}: producto no encontrado')
            continue

        precio = variante.precio_web
        if 'precio_unitario' in item:
            try:
                precio = Decimal(str(item['precio_unitario']))
            except Exception:
                pass

        stock_actual = InventarioService.obtener_stock_actual(variante, deposito)
        if stock_actual < int(cantidad):
            errores.append(
                f'{variante.nombre_completo}: stock insuficiente (hay {stock_actual}, pediste {cantidad})'
            )
            continue

        items_para_venta.append({
            'variante': variante,
            'cantidad': int(cantidad),
            'precio_unitario': precio,
            'descuento_unitario': Decimal('0'),
        })

    if errores:
        return Response(
            {'error': 'Errores de validación', 'detalle': errores},
            status=status.HTTP_400_BAD_REQUEST
        )

    if not items_para_venta:
        return Response(
            {'error': 'No hay ítems válidos para procesar'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Mapear método de pago del frontend al enum de Venta
    _METODO_MAP = {
        'mercadopago':   Venta.MetodoPago.MERCADOPAGO,
        'transferencia': Venta.MetodoPago.TRANSFERENCIA,
        'efectivo':      Venta.MetodoPago.EFECTIVO,
    }
    metodo_pago_raw = (data.get('metodo_pago') or 'transferencia').lower()
    metodo_pago = _METODO_MAP.get(metodo_pago_raw, Venta.MetodoPago.TRANSFERENCIA)

    try:
        with transaction.atomic():
            subtotal = sum(
                (it['precio_unitario'] - it['descuento_unitario']) * it['cantidad']
                for it in items_para_venta
            )
            # Aplicar costo de envío simple si corresponde
            total = subtotal + costo_envio
            venta = Venta.objects.create(
                cliente=None,
                usuario=usuario,
                deposito=deposito,
                subtotal=subtotal,
                descuento_porcentaje=0,
                descuento_monto=0,
                total=total,
                metodo_pago=metodo_pago,
                estado=Venta.EstadoVenta.PENDIENTE_PAGO,
            )

            for it in items_para_venta:
                st = (it['precio_unitario'] - it['descuento_unitario']) * it['cantidad']
                DetalleVenta.objects.create(
                    venta=venta,
                    variante=it['variante'],
                    cantidad=it['cantidad'],
                    precio_unitario=it['precio_unitario'],
                    descuento_unitario=it['descuento_unitario'],
                    subtotal=st,
                    costo_unitario=it['variante'].costo,
                )
                InventarioService.registrar_movimiento(
                    variante=it['variante'],
                    deposito=deposito,
                    tipo_movimiento=MovimientoStock.TipoMovimiento.VENTA_WEB,
                    cantidad=-it['cantidad'],
                    usuario=usuario,
                    referencia_tipo='venta',
                    referencia_id=venta.id,
                    observaciones=f'Venta web #{venta.numero}. {observaciones}',
                )

        # Vincular al ClienteWeb si está autenticado
        cliente_web = get_cliente_from_request(request)
        PedidoWeb.objects.create(venta=venta, cliente_web=cliente_web)

        # Enviar emails (no bloquea si falla)
        _enviar_emails_pedido(
            venta=venta,
            datos_cliente=datos_cliente,
            items_para_venta=items_para_venta,
            tipo_entrega=tipo_entrega,
            direccion_envio=direccion_envio or '',
        )

        return Response({
            'ok': True,
            'venta_id': venta.id,
            'venta_numero': venta.numero,
            'total': format(venta.total, '.2f'),
            'estado': venta.estado,
            'metodo_pago': venta.metodo_pago,
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        return Response(
            {'error': 'Error al crear el pedido', 'detalle': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([AllowAny])
def pedido_status(request, numero):
    """
    Estado público de un pedido web (por número de venta).
    No expone datos sensibles — solo número, estado, total y método de pago.
    """
    pedido_web = PedidoWeb.objects.filter(venta__numero=numero).select_related('venta').first()
    if not pedido_web:
        return Response({'error': 'Pedido no encontrado'}, status=status.HTTP_404_NOT_FOUND)
    v = pedido_web.venta
    return Response({
        'numero': v.numero,
        'estado': v.estado,
        'metodo_pago': v.metodo_pago,
        'total': format(v.total, '.2f'),
        'fecha': v.fecha.isoformat(),
    })


# --- Admin (requiere autenticación) ---


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_pedidos_list(request):
    """
    Lista pedidos de la tienda web (ventas con movimiento VENTA_WEB).
    Requiere login.
    """
    from apps.ventas.serializers import VentaListSerializer

    ids_web = MovimientoStock.objects.filter(
        tipo=MovimientoStock.TipoMovimiento.VENTA_WEB,
        referencia_tipo='venta',
    ).values_list('referencia_id', flat=True).distinct()

    qs = Venta.objects.filter(
        id__in=ids_web,
    ).select_related('cliente', 'usuario', 'deposito').prefetch_related('detalles__variante').order_by('-fecha')

    page_size = min(int(request.query_params.get('page_size', 20)), 100)
    paginator = Paginator(qs, page_size)
    page_num = request.query_params.get('page', 1)
    page = paginator.get_page(page_num)

    serializer = VentaListSerializer(page.object_list, many=True)
    return Response({
        'count': paginator.count,
        'total_pages': paginator.num_pages,
        'current_page': page.number,
        'results': serializer.data,
    })


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def admin_pedido_estado(request, pk):
    """
    Cambia el estado de un pedido web y notifica al cliente por email.
    Body: { estado: 'PAGO_CONFIRMADO' | 'EN_PREPARACION' | 'ENVIADO' | 'ENTREGADO' | 'ANULADA' }
    """
    ESTADOS_VALIDOS = {
        Venta.EstadoVenta.PAGO_CONFIRMADO,
        Venta.EstadoVenta.EN_PREPARACION,
        Venta.EstadoVenta.ENVIADO,
        Venta.EstadoVenta.ENTREGADO,
        Venta.EstadoVenta.ANULADA,
    }

    venta = Venta.objects.filter(pk=pk).first()
    if not venta:
        return Response({'error': 'Pedido no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    if not MovimientoStock.objects.filter(
        referencia_id=pk, tipo=MovimientoStock.TipoMovimiento.VENTA_WEB, referencia_tipo='venta'
    ).exists():
        return Response({'error': 'No es un pedido web'}, status=status.HTTP_400_BAD_REQUEST)

    nuevo_estado = (request.data.get('estado') or '').strip()
    if nuevo_estado not in ESTADOS_VALIDOS:
        return Response(
            {'error': f'Estado inválido. Opciones: {", ".join(ESTADOS_VALIDOS)}'},
            status=status.HTTP_400_BAD_REQUEST
        )

    venta.estado = nuevo_estado
    venta.save(update_fields=['estado'])

    # Notificar al cliente por email
    pedido_web = PedidoWeb.objects.filter(venta=venta).select_related('cliente_web').first()
    email_cliente = ''
    nombre_cliente = ''
    if pedido_web:
        # Extraer email del campo observaciones (datos_cliente guardados allí)
        obs = venta.detalles.first()  # no tiene email directo en Venta
        if pedido_web.cliente_web:
            email_cliente  = pedido_web.cliente_web.email
            nombre_cliente = pedido_web.cliente_web.nombre

    _enviar_email_estado(venta, email_cliente, nombre_cliente)

    return Response({
        'ok': True,
        'venta_id': venta.id,
        'numero': venta.numero,
        'estado': venta.estado,
        'estado_label': _ESTADO_LABELS.get(venta.estado, venta.estado),
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_pedido_detail(request, pk):
    """Detalle de pedido web. Requiere login."""
    from apps.ventas.serializers import VentaSerializer

    venta = Venta.objects.select_related(
        'cliente', 'usuario', 'deposito'
    ).prefetch_related('detalles__variante', 'detalles__variante__producto_base').filter(pk=pk).first()

    if not venta:
        return Response({'error': 'Pedido no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    # Verificar que es pedido web
    if not MovimientoStock.objects.filter(
        referencia_id=pk,
        tipo=MovimientoStock.TipoMovimiento.VENTA_WEB,
        referencia_tipo='venta',
    ).exists():
        return Response({'error': 'Pedido no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    serializer = VentaSerializer(venta)
    return Response(serializer.data)


# ============================================================================ #
# Auth de clientes web                                                           #
# ============================================================================ #

def _cliente_dict(cliente: ClienteWeb) -> dict:
    return {
        'id': cliente.id,
        'nombre': cliente.nombre,
        'email': cliente.email,
        'avatar_url': cliente.avatar_url,
        'fecha_registro': cliente.fecha_registro.isoformat(),
    }


@api_view(['POST'])
@permission_classes([AllowAny])
def auth_registro(request):
    """
    Registro con email y contraseña.
    Body: { nombre, email, password }
    """
    data = request.data or {}
    nombre = (data.get('nombre') or '').strip()
    email = (data.get('email') or '').strip().lower()
    password = data.get('password') or ''

    if not nombre or not email or not password:
        return Response({'error': 'nombre, email y password son obligatorios'}, status=400)
    if len(password) < 6:
        return Response({'error': 'La contraseña debe tener al menos 6 caracteres'}, status=400)
    if ClienteWeb.objects.filter(email=email).exists():
        return Response({'error': 'Ya existe una cuenta con ese email'}, status=400)

    cliente = ClienteWeb.objects.create(
        nombre=nombre,
        email=email,
        password_hash=hash_password(password),
    )
    return Response({
        'token': generate_token(cliente),
        'user': _cliente_dict(cliente),
    }, status=201)


@api_view(['POST'])
@permission_classes([AllowAny])
def auth_login(request):
    """
    Login con email y contraseña.
    Body: { email, password }
    """
    data = request.data or {}
    email = (data.get('email') or '').strip().lower()
    password = data.get('password') or ''

    try:
        cliente = ClienteWeb.objects.get(email=email, activo=True)
    except ClienteWeb.DoesNotExist:
        return Response({'error': 'Email o contraseña incorrectos'}, status=401)

    if not cliente.password_hash or not verify_password(password, cliente.password_hash):
        return Response({'error': 'Email o contraseña incorrectos'}, status=401)

    return Response({
        'token': generate_token(cliente),
        'user': _cliente_dict(cliente),
    })


@api_view(['POST'])
@permission_classes([AllowAny])
def auth_google(request):
    """
    Login / registro con Google.
    Body: { id_token: <Google ID token del frontend> }
    """
    data = request.data or {}
    id_token = data.get('id_token') or ''
    if not id_token:
        return Response({'error': 'id_token es obligatorio'}, status=400)

    payload = verify_google_token(id_token)
    if not payload:
        return Response({'error': 'Token de Google inválido'}, status=401)

    google_id = payload.get('sub') or ''
    email = (payload.get('email') or '').lower()
    nombre = payload.get('name') or email.split('@')[0]
    avatar = payload.get('picture') or ''

    if not email:
        return Response({'error': 'No se pudo obtener el email de Google'}, status=400)

    # Buscar por google_id o email (puede ser un usuario que ya se registró con email)
    cliente = (
        ClienteWeb.objects.filter(google_id=google_id).first()
        or ClienteWeb.objects.filter(email=email).first()
    )

    if cliente:
        # Actualizar datos de Google si es la primera vez con Google
        update_fields = []
        if not cliente.google_id:
            cliente.google_id = google_id
            update_fields.append('google_id')
        if avatar and not cliente.avatar_url:
            cliente.avatar_url = avatar
            update_fields.append('avatar_url')
        if update_fields:
            cliente.save(update_fields=update_fields)
    else:
        cliente = ClienteWeb.objects.create(
            nombre=nombre,
            email=email,
            google_id=google_id,
            avatar_url=avatar,
        )

    return Response({
        'token': generate_token(cliente),
        'user': _cliente_dict(cliente),
    })


@api_view(['GET'])
@permission_classes([AllowAny])
def auth_me(request):
    """Devuelve los datos del cliente autenticado."""
    cliente = get_cliente_from_request(request)
    if not cliente:
        return Response({'error': 'No autenticado'}, status=401)
    return Response(_cliente_dict(cliente))


@api_view(['GET'])
@permission_classes([AllowAny])
def mis_pedidos(request):
    """Historial de pedidos del cliente autenticado."""
    cliente = get_cliente_from_request(request)
    if not cliente:
        return Response({'error': 'No autenticado'}, status=401)

    pedidos_web = (
        PedidoWeb.objects
        .filter(cliente_web=cliente)
        .select_related('venta')
        .prefetch_related('venta__detalles__variante')
        .order_by('-fecha')
    )

    results = []
    for pw in pedidos_web:
        v = pw.venta
        items = []
        for det in v.detalles.all():
            nombre_v = getattr(det.variante, 'nombre_completo', str(det.variante))
            items.append({
                'nombre': nombre_v,
                'cantidad': det.cantidad,
                'precio_unitario': str(det.precio_unitario),
                'subtotal': str(det.subtotal),
            })
        results.append({
            'id': pw.id,
            'venta_numero': v.numero,
            'fecha': pw.fecha.isoformat(),
            'total': str(v.total),
            'estado': v.estado,
            'items': items,
        })

    return Response(results)
